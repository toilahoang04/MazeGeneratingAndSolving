import pygame as pg
import sys
import random
import copy
import heapq
from collections import deque
import pandas as pd
import pygame as pg
import sys
from openpyxl import Workbook, load_workbook

pg.init()
# size
mazeSize=500
buttonbarW=250
barH=500
pointbarW=150
border=10
screenW=border+mazeSize+buttonbarW+pointbarW
screenH=border+barH
# Các màu cơ bản
white = (255, 255, 255)         # Trắng
black = (0, 0, 0)                # Đen
gray = (128, 128, 128)           # Xám
red = (255, 0, 0)                # Đỏ
green = (0, 255, 0)              # Xanh lá
blue = (0, 0, 255)               # Xanh dương
aqua = (0, 255,255) 
dark_green = (0, 100, 0)
pale_yellow = (255, 255, 204)
displaySurf=pg.display.set_mode((border+mazeSize+buttonbarW+pointbarW,border+barH))
displaySurf.fill(gray)

#Ảnh
startPic=pg.image.load("./img/actor.png")
endPic=pg.image.load("./img/vietnam.png")
bg=pg.image.load("./img/bg.png")
aboutUs=pg.image.load("./img/AboutUs.png")
left=pg.image.load("./img/left.png")
right=pg.image.load("./img/right.png")
#scale ảnh
startPic=pg.transform.scale(startPic,(20,20))
endPic=pg.transform.scale(endPic,(20,20))
bg=pg.transform.scale(bg,(screenW,screenH))
left=pg.transform.scale(left,(20,20))
right=pg.transform.scale(right,(20,20))

#load music
pg.mixer.music.load('./DLTTAD.mp3')  # Thay thế 'background_music.mp3' bằng tên tệp của bạn
pg.mixer.music.set_volume(0.01)  # Thiết lập âm lượng (0.0 đến 1.0)
pg.mixer.music.play(-1)  # Phát nhạc nền lặp lại vô hạn (-1)
#volume = 0.5



#function chức năng
def drawButton(x,y,w,h,content,size,textColor,inactiveColor,activeColor):
    mouse = pg.mouse.get_pos()
    if x + w > mouse[0] > x and y + h > mouse[1] > y:
        pg.draw.rect(displaySurf, activeColor, (x, y, w, h))
    else:
        pg.draw.rect(displaySurf, inactiveColor, (x, y, w, h))
    font=pg.font.Font(None,size)
    text=font.render(content,True,textColor)
    textRect=text.get_rect()
    textBlockRect=pg.Rect(x,y,w,h)
    textRect.center=textBlockRect.center
    displaySurf.blit(text,textRect)
def drawRecAndText(x,y,w,h,content,size,boxColor,textColor):
    pg.draw.rect(displaySurf, boxColor, (x, y, w, h))
    font=pg.font.Font(None,size)
    text=font.render(content,True,textColor)
    textRect=text.get_rect()
    textBlockRect=pg.Rect(x,y,w,h)
    textRect.center=textBlockRect.center
    displaySurf.blit(text,textRect)

def drawText(x,y,w,h,content,size,color):
    font=pg.font.Font(None,size)
    text=font.render(content,True,color)
    textRect=text.get_rect()
    textBlockRect=pg.Rect(x,y,w,h)
    textRect.center=textBlockRect.center
    displaySurf.blit(text,textRect)
#Biến check đang ở screen nào
HomeScreen=True
AiScreen=False
WinScreen=False
AboutUsScreen=False
HistoryScreen = False
# Tên file Excel
file_name = "History Algorithm.xlsx"
# Biến scroll
scroll_y = 0  # Vị trí cuộn theo trục Y
SCROLL_SPEED = 20  # Tốc độ cuộn
#check map
haveMap=False
#check block đang cầm
inhand=0 #0 rỗng; 1 Tường; 2 Đường
#hàm vẽ screen
start=True
# numOfMap số lượng map đang có, indexMap map hiện tại
with open('map/NOM.txt', 'r', encoding='utf-8') as file:
    content = file.read().strip()  # loại bỏ khoảng trắng và ký tự xuống dòng
    numOfMap = int(content)
    print(numOfMap)
def ReadNOM():
    with open('map/NOM.txt', 'r', encoding='utf-8') as file:
        content = file.read().strip()
        n = int(content)
        return n
def UpdateNOM():
    global numOfMap
    with open('map/NOM.txt', 'w', encoding='utf-8') as file:
        file.write(str(numOfMap) + '\n')
numOfMap=ReadNOM()
if numOfMap>1:
    indexMap=1
else: 
    indexMap=0

def WriteMaze(matrix):
    global numOfMap
    numOfMap=numOfMap+1
    UpdateNOM()
    with open(f"map/maze{numOfMap}.txt", 'w', encoding='utf-8') as file:
        for row in matrix:
            line = ' '.join(str(element) for element in row)
            file.write(line + '\n')
def ReadMaze():
    global indexMap
    matrix = []
    with open(f"map/maze{indexMap}.txt", 'r', encoding='utf-8') as file:
        for line in file:
            row = [float(num) for num in line.strip().split()]
            matrix.append(row)
    return matrix
def drawAiScreen():
    global haveMap
    global inhand
    global indexMap
    if not haveMap:
        displaySurf.blit(bg,(0,0))
        haveMap=True
    pg.draw.rect(displaySurf,(0,0,255),(border/2,border/2,buttonbarW,barH))
    #pg.draw.rect(displaySurf,(255,255,255),(border/2+buttonbarW,border/2,mazeSize,mazeSize))
    pg.draw.rect(displaySurf,(0,255,0),(border/2+mazeSize+buttonbarW,border/2,pointbarW,barH))
    #left bar
    drawRecAndText(border/2,border/2*1+50*0,buttonbarW,50,'Autogenerate Maze',24,black,white)


    drawButton(border,border/2*2+50*1,buttonbarW-border,50,'DFS',24,black,white,red)
    drawButton(border,border/2*3+50*2,buttonbarW-border,50,'Hunt and Kill',24,black,white,red)

    drawRecAndText(border/2,border/2*4+50*3,buttonbarW,50,'Algorithm Solving Maze',24,black,white)


    drawButton(border,border/2*5+50*4,buttonbarW-border,50,'BFS',24,black,white,red)
    drawButton(border,border/2*6+50*5,buttonbarW-border,50,'DFS',24,black,white,red)
    drawButton(border,border/2*7+50*6,buttonbarW-border,50,'Greedy',24,black,white,red)
    drawButton(border,border/2*8+50*7,buttonbarW-border,50,'A*',24,black,white,red)
    drawButton(border,border/2*9+50*8,buttonbarW-border,50,'Local Search',24,black,white,red)
    #right bar
    drawText(border+buttonbarW+mazeSize,border/2+50*0,75,50,'Start point',24,red)
    drawText(border+buttonbarW+mazeSize,border/2+50*1,75,50,'End point',24,red)
    displaySurf.blit(startPic,(border+buttonbarW+mazeSize+75+10,border/2+15))
    displaySurf.blit(endPic,(border+buttonbarW+mazeSize+75+10,border/2+15*4))

    drawButton(border+buttonbarW+mazeSize,border/2+50*2,140,40,'User can play',24,black,white,red)
    drawButton(border+buttonbarW+mazeSize,border/2+50*3,140,40,'Back to home',24,black,white,red)
    #chọn, lưu map===
    displaySurf.blit(left,(border+buttonbarW+mazeSize,border/2+50*4))
    displaySurf.blit(right,(border+buttonbarW+mazeSize+50,border/2+50*4))
    drawButton(border+buttonbarW+mazeSize+20,border/2+50*4,30,20,f"{indexMap}",16,black,white,white)
    drawButton(border+buttonbarW+mazeSize+70,border/2+50*4,70,20,'Choice',16,black,white,red)
    drawButton(border+buttonbarW+mazeSize+70,border/2+50*4+25,70,20,'Save',16,black,white,red)
    #================
    pg.draw.rect(displaySurf,white,(border+buttonbarW+mazeSize,border/2+50*5,140,210))
    drawText(border+buttonbarW+mazeSize+10,border/2+50*5+10,120,30,"Edit Maze",28,black)
    drawButton(border+buttonbarW+mazeSize+10,border/2+50*5+40*1,120,30,'Get wall',16,white,black,red)
    drawButton(border+buttonbarW+mazeSize+10,border/2+50*5+40*2,120,30,'Get empty block',16,white,black,red)
    drawButton(border+buttonbarW+mazeSize+10,border/2+50*5+40*3,120,30,'Clean the maze',16,white,black,red)
    drawText(border+buttonbarW+mazeSize+10,border/2+50*5+40*4,120,30,"You're holding:",24,black)
    if inhand==0:
        drawText(border+buttonbarW+mazeSize+10,border/2+50*5+40*4+20,120,30,"nothing",24,red)
    elif inhand==1:
        drawText(border+buttonbarW+mazeSize+10,border/2+50*5+40*4+20,120,30,"wall",24,red)
    elif inhand==2:
        drawText(border+buttonbarW+mazeSize+10,border/2+50*5+40*4+20,120,30,"empty block",24,red)
def drawHomeScreen():
    displaySurf.blit(bg,(0,0))
    drawText(screenW/2-150,50,300,300,"MAZE GAME",100,red)
    drawButton(screenW/2-100,250,200,50,"Play",16,white,black,red)
    drawButton(screenW/2-100,325,200,50,"Creator Info",16,white,black,red)
    drawButton(screenW/2-100,400,200,50,"Saved data",16,white,black,red)
def drawAboutUsScreen():
    displaySurf.blit(bg,(0,0))
    displaySurf.blit(aboutUs,(60,10))
    drawButton(screenW/2-100,400,200,50,"Home",16,white,black,red)
def drawWinScreen():
    displaySurf.blit(bg,(0,0))
    drawText(screenW/2-150,50,300,300,"You Win",100,red)
    drawButton(screenW/2-100,250,200,50,"Come back",16,white,black,red)
def drawHistoryScreen():
    displaySurf.blit(bg,(0,0))
    drawButton(screenW/2-350,50,200,50,"back to Home",16,white,black,red)
def Win():
    global AiScreen
    global WinScreen
    print("Win")
    AiScreen=False
    WinScreen=True

# Kiểm tra nếu file chưa tồn tại, tạo mới
try:
    wb = load_workbook(file_name)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    # Thêm tiêu đề cột
    # Định nghĩa các tên cột
    columns = [
    "STT",
    "Algorithm",
    "Start position",
    "End position",
    "Total steps",
    "Result steps" 
    ]
    ws.append(columns)
dataTable = 0
tableWidth = 0
tableHeight = 0
wb.save(file_name) 
df = pd.read_excel(file_name)
lenTable = [df.columns.tolist()] + df.values.tolist()  # Chuyển thành danh sách 2D
count = len(lenTable) - 1
def saveRowsData(data):
    global dataTable,tableWidth ,tableHeight
    ws.append(data)
    wb.save(file_name)  # Lưu thay đổi sau mỗi lần ghi
    df = pd.read_excel(file_name)
    dataTable = [df.columns.tolist()] + df.values.tolist()  # Chuyển thành danh sách 2D
    #print(dataTable,"88888888")
    
    tableWidth = len(dataTable[0]) * 200
    tableHeight = len(dataTable) * 50

def drawTable(surface, data, start_x, start_y,scroll_y,cell_width, cell_height):
   # Font chữ
    font = pg.font.Font(None, 30)
    #print(data)
    drawText(450,50-scroll_y,50,50,"History Algorithm",50,white)
    for row_idx, row in enumerate(data):
        y = start_y + row_idx * cell_height - scroll_y
        if y + cell_height < 0 or y > screenH:  # Bỏ qua nếu ngoài màn hình
            continue
        for col_idx, cell in enumerate(row):
            x = start_x + col_idx * cell_width
            # Vẽ ô
            pg.draw.rect(surface, white, (x, y, cell_width, cell_height), 1)
            # Vẽ nội dung
            if col_idx == 4: 
                text_surface = font.render(str(cell), True, red)
            elif col_idx == 5:
                text_surface = font.render(str(cell), True, aqua)
            else:
                text_surface = font.render(str(cell), True, white)
            text_rect = text_surface.get_rect(center=(x + cell_width // 2, y + cell_height // 2))
            surface.blit(text_surface, text_rect)
     
#saveRowsData([count,"A*", "(0, 0)", "(5, 5)", 15, 10])


def check_stable_button(x, y, width, height):
    mouse = pg.mouse.get_pos()
    click = pg.mouse.get_pressed()
    if x + width > mouse[0] > x and y + height > mouse[1] > y:
        if click[0] == 1:
            return True
    return False

def EditMaze(x, y, width, height,Maze):
    mouse = pg.mouse.get_pos()
    click = pg.mouse.get_pressed()
    if x + width > mouse[0] > x and y + height > mouse[1] > y:
        if click[0] == 1:
            cellx=int((mouse[0]-border/2-buttonbarW)//20)
            celly=int((mouse[1]-border/2)//20)
            print(cellx)
            print(celly)
            if inhand==1 and Maze[cellx][celly]==0:
                Maze[cellx][celly]=1
            elif inhand==2 and Maze[cellx][celly]==1:
                Maze[cellx][celly]=0
            return True
    return False


# check button dfs
check_button_DFS_maze = False
check_button_Hunt_and_Kill_maze = False
check_button_dfs = False
check_button_bfs = False
check_button_Greedy = False
check_button_Astar = False
check_button_localsearch = False
#check người chơi
check_user_play = False

# Directions for moving in the maze (Right, Left, Down, Up)
# Khởi tạo các hướng di chuyển
DIRECTIONS = [
    (-1, 0),  # Lên
    (1, 0),   # Xuống
    (0, -1),  # Trái
    (0, 1)    # Phải
]

cell_size = 20
w,h = mazeSize // cell_size, mazeSize // cell_size



SaveEndx,SaveEndy =0,0
SaveStartx,SaveStarty =0,0
MAZE = [[0 for _ in range(25)] for _ in range(25)]
for i in range(25):
    for j in range(25):
        if i==0 or i==24 or j==0 or j==24:
            MAZE[i][j]=1
        else:
            MAZE[i][j]=0
startx = (border/2+buttonbarW) 
starty = (border/2) 
#print(maze)
#=====================================Function chung========================================
# Hàm kiểm tra xem vị trí mới có nằm trong phạm vi của mê cung và là tường không
def is_within_bounds(x, y):
    return 0 < x < h-1 and 0 < y < w-1 and maze[x][y] == 1
#==========================================DFS==============================================
# Randomized Depth-First Search algorithm to generate maze
def generate_maze(x, y,maze):
    x = int(x)
    y = int(y)
    maze[x][y] = 0  # Đặt ô khởi đầu là đường (0)
    random.shuffle(DIRECTIONS)  # Trộn hướng đi để tạo mê cung ngẫu nhiên

    for dx, dy in DIRECTIONS:
        nx, ny = x + dx * 2, y + dy * 2  # Di chuyển hai ô mỗi lần để tạo tường ngăn cách
        if is_within_bounds(nx, ny):
            maze[x + dx][y + dy] = 0  # Xóa tường giữa các ô
            maze[nx][ny] = 0          # Xóa ô đích để tạo đường đi
            draw_maze(startx,starty,maze)               # Cập nhật màn hình mỗi khi có thay đổi
            pg.display.flip()     # Vẽ lại màn hình
            pg.time.delay(20)     # Thời gian trễ để xem rõ quá trình tạo mê cung
            generate_maze(nx, ny,maze)     # Đệ quy tiếp tục từ ô đích
# Drawing the maze
def draw_maze(startx, starty,maze):
    for x in range(w):
        for y in range(h):
            if maze[x][y] == 1 :
                color = dark_green
            elif maze[x][y] == 5:
                color = blue
            elif maze[x][y] == 4:
                color = red
            elif maze[x][y] == 3:
                color = aqua
            else: color = white
            pg.draw.rect(displaySurf, color, (startx + x * cell_size, starty+ y * cell_size, cell_size, cell_size))
draw_maze(startx,starty,MAZE)
def drawStartEnd():
    list =[]
    for x in range(w):
        for y in range(h):
            if maze[x][y] == 0 :
                list.append((x,y))
    randomend = random.randint(0, len(list))
    #print(randomend)
    tuple = list[randomend]
    #print(tuple)
    
    pg.draw.rect(displaySurf, red, (startx + randomstartx * cell_size, starty+ randomstarty* cell_size, cell_size, cell_size))
    pg.draw.rect(displaySurf, aqua, (startx + tuple[0] * cell_size, starty+ tuple[1]* cell_size, cell_size, cell_size))
    return tuple
    #print(startx , randomstartx, starty, randomstarty)

#==========================================Hunt and kill==============================================
def createMaze(maze,n):
    for i in range(n):
        for j in range(n):
            if i%2==0 or j%2==0:
                maze[i][j]=1
            else:
                maze[i][j]=0
def is_valid(maze, x, y):
    # Kiểm tra xem ô (x, y) có nằm trong giới hạn của mê cung và chưa được ghé thăm
    rows, cols = len(maze), len(maze[0])
    return 0 <= x < rows and 0 <= y < cols and maze[x][y] == 0
def huntAndKill(maze, start_x, start_y):
    maze[start_x][start_y] = 2  # Đánh dấu ô đầu tiên là đã ghé thăm bằng 2
    current_x, current_y = start_x, start_y
    while True:
        # Kill phase: Tìm ô lân cận chưa ghé thăm
        unvisited_neighbors = []
        for dx, dy in DIRECTIONS:
            nx, ny = current_x + 2*dx, current_y + 2*dy
            if is_valid(maze, nx, ny):
                unvisited_neighbors.append((nx, ny))

        if unvisited_neighbors:
            # Chọn ngẫu nhiên một ô lân cận chưa ghé thăm
            next_x, next_y = random.choice(unvisited_neighbors)
            wall_x, wall_y = (current_x + next_x) // 2, (current_y + next_y) // 2
            maze[wall_x][wall_y] = 2  # Phá vách ngăn bằng cách đánh dấu là 2
            maze[next_x][next_y] = 2  # Đánh dấu ô mới là đã ghé thăm
            current_x, current_y = next_x, next_y
            #vẽ
            draw_maze(startx,starty,maze)
            pg.display.flip()
            pg.time.delay(20)
        else:
            # Hunt phase: Tìm ô chưa ghé thăm có lân cận đã ghé thăm
            found = False
            for i in range(len(maze)):
                for j in range(len(maze[i])):
                    if maze[i][j] == 0:  # Tìm ô chưa ghé thăm
                        # Kiểm tra nếu ô này có ô lân cận đã ghé thăm
                        for dx, dy in DIRECTIONS:
                            ni, nj = i + 2*dx, j + 2*dy
                            if 0 <= ni < len(maze) and 0 <= nj < len(maze[0]) and maze[ni][nj] == 2:
                                # Phá vách ngăn và đánh dấu ô mới là đã ghé thăm
                                wall_x, wall_y = (i + ni) // 2, (j + nj) // 2
                                maze[wall_x][wall_y] = 2  # Phá vách ngăn
                                maze[i][j] = 2  # Đánh dấu ô này là đã ghé thăm
                                current_x, current_y = i, j  # Cập nhật vị trí hiện tại
                                found = True
                                pg.time.delay(50)
                                break
                        if found:
                            break
                    if found:
                        break
            # Nếu không tìm thấy ô nào chưa ghé thăm, thoát khỏi vòng lặp
            if not found:
                break

    # In mê cung kết quả
    # for row in maze:
    #     print(" ".join(map(str, row)))
draggingStart = False
draggingEnd = False


# startPicSmall = pg.transform.scale(startPic,(40,40))
# endPicSmall = pg.transform.scale(endPic,(40,40))
imgStart = startPic.get_rect()
imgEnd = endPic.get_rect()



# creatr start point and end point
checkCreateStartEndDFS = False
checkCreateStartEndHuntAndKill=False
beforexStart = 0
beforeyStart = 0
beforexEnd = 0
beforeyEnd = 0

#hàm vẽ thông báo
def drawWin():
    displaySurf.blit(bg,(0,0))
    drawText(screenW/2-150,screenH/2-150,300,300,"A path is available",100,red)
    pg.display.update()
    pg.time.delay(2000)
def drawLose():
    displaySurf.blit(bg,(0,0))
    drawText(screenW/2-150,screenH/2-150,300,300,"No path is available",100,red)
    pg.display.update()
    pg.time.delay(2000)



# Find position of start and end

def findStartEnd(maze):
    Start = ()
    End = ()
    for x in range(w):
        for y in range(h):
            if maze[x][y] == 4:
                Start = (x, y)
            if maze[x][y] == 3:
                End = (x, y)
    return Start, End
# def solve maze
checkCreateMap = True
checkCreateStartPoint = False
checkCreateEndPoint = False 
def finddfs(maze, start, goal):
    # Khởi tạo stack và tập các ô đã đi qua
    stack = []
    stack.append(start)
    visited = set()
    visited.add(start)
    
    # Để lưu lại đường đi
    came_from = {}  # Dùng để truy vết lại đường đi đã qua
    came_from[start] = None  # Điểm bắt đầu không có ô trước nó
    
    # Khởi tạo các hướng di chuyển
    directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]  # lên, xuống, trái, phải
    
    # Duyệt stack
    while stack:
        current = stack.pop()
        x, y = current
        
        # Kiểm tra nếu là đích
        if maze[x][y] == 3:
            # Truy vết lại từ đích về điểm bắt đầu
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()  # Đảo ngược để có thứ tự từ bắt đầu đến đích
            return path  # Trả về đường đi đã tìm được
        elif start != current:
            maze[x][y] = 5
            drawCell(x, y, red)
            pg.time.delay(50)
            pg.display.update()
        
        # Kiểm tra các ô lân cận
        for dx, dy in directions:
            nx, ny = x + dx, y + dy
            
            # Kiểm tra tính hợp lệ của ô lân cận
            if (0 <= nx < len(maze) and 0 <= ny < len(maze[0]) and
                (nx, ny) not in visited and maze[nx][ny] != 1):
                
                # Thêm ô mới vào stack
                stack.append((nx, ny))
                visited.add((nx, ny))
                
                # Lưu lại đường đi
                came_from[(nx, ny)] = current
    
    return None  # Không tìm thấy đường đi

def findbfs(maze,start,end):
    # Khởi tạo queue và tập các ô đã đi qua
    queue = deque()
    queue.append(start)
    visited = set()
    visited.add(start)
    
    # Để lưu lại đường đi
    came_from = {}  # Dùng để truy vết lại đường đi đã qua
    came_from[start] = None  # Điểm bắt đầu không có ô trước nó
    
    # Khởi tạo các hướng di chuyển
    directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]  # lên, xuống, trái, phải
    
    # Duyệt queue
    while queue:
        current = queue.popleft()
        x, y = current
        
        # Kiểm tra nếu là đích
        if maze[x][y] == 3:
            # Truy vết lại từ đích về điểm bắt đầu
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()  # Đảo ngược để có thứ tự từ bắt đầu đến đích
            return path  # Trả về đường đi đã tìm được
        elif start != current:
            maze[x][y] = 5
            drawCell(x, y, red)
            pg.time.delay(50)
            pg.display.update()
        
        # Kiểm tra các ô lân cận
        for dx, dy in directions:
            nx, ny = x + dx, y + dy
            
            # Kiểm tra tính hợp lệ của ô lân cận
            if (0 <= nx < len(maze) and 0 <= ny < len(maze[0]) and
                (nx, ny) not in visited and maze[nx][ny] != 1):
                
                # Thêm ô mới vào queue
                queue.append((nx, ny))
                visited.add((nx, ny))
                
                # Lưu lại đường đi
                came_from[(nx, ny)] = current
    
    return None  # Không tìm thấy đường đi
    
   
        
# ==============================thuật toán tìm đường có hướng====================
# hàm chung
def manhattan_distance(x1, y1, x2, y2):
    return abs(x1 - x2) + abs(y1 - y2)
def drawCell(x,y,color): #hàm tô màu 1 ô
    pg.draw.rect(displaySurf,color,(border/2+buttonbarW+20*x,border/2+20*y,20,20))
def CleanWrongWay(maze):
    for i in range(w):
        for j in range(h):
            if(maze[i][j]==5):
                maze[i][j]=0
                drawCell(i,j,white)
def drawPic(pic,x,y):
    displaySurf.blit(pic,(border/2+buttonbarW+20*x,border/2+20*y))
    #displaySurf.blit(endPic,(startx + end[0]*cell_size,starty + end[1]*cell_size))
# Greedy
def Moving(Maze, Start, End, path):
    CleanWrongWay(Maze)
    if path:
        print("path is:")
        for step in path:
            Maze[step[0]][step[1]]=5
            print(f"{step[0]}; {step[1]}")
            drawCell(step[0],step[1],aqua)
        drawPic(startPic,Start[0],Start[1])
        drawPic(endPic,End[0],End[1])
        pg.display.update()
        pg.time.delay(1000)
        x=0
        y=0
        for step in path:
            drawPic(startPic,step[0],step[1])
            if x!=0 and y!=0:
                drawCell(x,y,white)
            x=step[0]
            y=step[1]
            pg.display.update()
            pg.time.delay(50)
        pg.time.delay(1000)
    else:
        print("can not find")
def findGreedy(maze, start, goal):
    # Khởi tạo hàng đợi ưu tiên và tập các ô đã đi qua
    queue = []
    heapq.heappush(queue, (0, start))  # (khoảng cách tới đích, vị trí)
    visited = set()
    visited.add(start)
    
    # Để lưu lại đường đi
    came_from = {}  # Dùng để truy vết lại đường đi đã qua
    came_from[start] = None  # Điểm bắt đầu không có ô trước nó
    
    # Khởi tạo các hướng di chuyển
    directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]  # lên, xuống, trái, phải
    
    # Duyệt hàng đợi
    while queue:
        _, current = heapq.heappop(queue)
        x, y = current
        # Kiểm tra nếu là đích
        if maze[x][y] == 3:
            # Truy vết lại từ đích về điểm bắt đầu
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()  # Đảo ngược để có thứ tự từ bắt đầu đến đích
            return path  # Trả về đường đi đã tìm được
        elif(start!=current):
            maze[x][y]=5
            drawCell(x,y,red)
            pg.time.delay(50)
            pg.display.update()
        
        # Kiểm tra các ô lân cận
        for dx, dy in directions:
            nx, ny = x + dx, y + dy
            
            # Kiểm tra tính hợp lệ của ô lân cận
            if (0 <= nx < len(maze) and 0 <= ny < len(maze[0]) and
                (nx, ny) not in visited and maze[nx][ny] != 1):
                
                # Thêm ô mới vào hàng đợi với độ ưu tiên dựa trên khoảng cách tới đích
                distance = manhattan_distance(nx, ny, goal[0], goal[1])
                heapq.heappush(queue, (distance, (nx, ny)))
                visited.add((nx, ny))
                
                # Lưu lại đường đi
                came_from[(nx, ny)] = current
    
    return None  # Không tìm thấy đường đi
# Astar
def findAstar(maze, start, goal):
    # Khởi tạo hàng đợi ưu tiên và các tập hợp dữ liệu
    queue = []
    heapq.heappush(queue, (0, start))  # (f(n), vị trí)
    g_cost = {start: 0}  # Chi phí g(n) từ điểm bắt đầu đến mỗi ô
    came_from = {start: None}  # Lưu vết đường đi

    # Các hướng di chuyển
    directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]  # lên, xuống, trái, phải

    # Vòng lặp chính
    while queue:
        _, current = heapq.heappop(queue)
        x, y = current
        # Kiểm tra nếu đã đến đích
        if maze[current[0]][current[1]] == 3:
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()  # Đảo ngược danh sách để có thứ tự từ start đến goal
            return path
        elif(start!=current):
            maze[x][y]=5
            drawCell(x,y,red)
            pg.time.delay(50)
            pg.display.update()
        # Duyệt các ô lân cận
        for dx, dy in directions:
            nx, ny = current[0] + dx, current[1] + dy
            next_cell = (nx, ny)
            
            # Kiểm tra tính hợp lệ của ô lân cận
            if (0 <= nx < len(maze) and 0 <= ny < len(maze[0]) and maze[nx][ny] != 1):
                # Tính chi phí g(n) cho ô lân cận
                new_cost = g_cost[current] + 1  # Giả sử mỗi bước có chi phí là 1

                # Nếu ô lân cận chưa được duyệt hoặc tìm thấy đường đi rẻ hơn đến ô này
                if next_cell not in g_cost or new_cost < g_cost[next_cell]:
                    g_cost[next_cell] = new_cost
                    f_cost = new_cost + manhattan_distance(nx, ny, goal[0], goal[1])  # f(n) = g(n) + h(n)
                    heapq.heappush(queue, (f_cost, next_cell))
                    came_from[next_cell] = current

    return None  # Không tìm thấy đường đi
def is_valid_move(maze, x, y):
    rows, cols = len(maze), len(maze[0])
    return 0 <= x < rows and 0 <= y < cols and (maze[x][y] == 0 or maze[x][y] == 4 or maze[x][y] == 3)
def findlocalsearch(maze,start,end):
    stack = [start]  # Ngăn xếp lưu các vị trí đã đi qua
    path = [start]  # Lưu đường đi đã chọn
    visited = set()
    visited.add(start)
    
    while stack:
        # print(stack[-1])
        current_position = stack[-1]  # Lấy vị trí hiện tại từ đỉnh ngăn xếp
        # if current_position is None:
        #     print("Lỗi: current_position không hợp lệ!")
        #     return None
        x, y = current_position
        neighbors = [
            (x + 1, y), (x - 1, y), 
            (x, y + 1), (x, y - 1)
        ]
        
        # Lọc các neighbor hợp lệ
        valid_neighbors = [
            (nx, ny) for nx, ny in neighbors 
            if is_valid_move(maze, nx, ny) and (nx, ny) not in visited
        ]
        
        # In tất cả các neighbor hợp lệ
        #print(f"Current position: {current_position}")
        #print(f"Valid neighbors: {valid_neighbors}")
        
        if not valid_neighbors:
            stack.pop()  
            path.pop()
            continue
        
        # Tìm neighbor gần nhất tới đích
        next_position = min(
            valid_neighbors, 
            key=lambda pos: manhattan_distance(pos[0], pos[1], end[0], end[1])
        )
        stack.append(next_position)
        visited.add(next_position)
        path.append(next_position)
        if next_position[0] == end[0] and next_position[1] == end[1] :
            return path
        else:
            maze[x][y] = 5
            drawCell(x, y, red)
            displaySurf.blit(startPic,(startx + start[0]*cell_size,starty + start[1]*cell_size))
            displaySurf.blit(endPic,(startx + end[0]*cell_size,starty + end[1]*cell_size))
            pg.time.delay(50)
            pg.display.update()
        
        #print(f"Next position chosen: {next_position}\n")    
    return None

def move(array,xStart,yStart,xEnd,yEnd,stable):
    if stable == 1 and xStart - 1 >= 0 and (array[xStart - 1][yStart] == 0 or array[xStart - 1][yStart] == 3) :
            
        array[xStart - 1][yStart] = 4 
        array[xStart][yStart] = 0
        if xStart-1 != xEnd or yStart != yEnd:
            array[xEnd][yEnd] = 3
        else:
            Win()
        draw_maze(startx,starty,array)
        displaySurf.blit(startPic,(startx + (xStart-1)*cell_size,starty + yStart*cell_size))
        displaySurf.blit(endPic,(startx + xEnd*cell_size,starty + yEnd*cell_size))
        
    
    elif stable == 2 and xStart + 1 < w and (array[xStart + 1][yStart] == 0  or array[xStart + 1][yStart] == 3):
        array[xStart + 1][yStart] = 4 
        array[xStart][yStart] = 0 
        if xStart+1 != xEnd or yStart != yEnd:
            array[xEnd][yEnd] = 3
        else:
            Win()
        draw_maze(startx,starty,array)    
        displaySurf.blit(startPic,(startx + (xStart+1)*cell_size,starty + yStart*cell_size))
        displaySurf.blit(endPic,(startx + xEnd*cell_size,starty + yEnd*cell_size))
      
    elif stable == 3 and yStart - 1 >= 0 and (array[xStart][yStart - 1] == 0 or array[xStart][yStart - 1] == 3):
        array[xStart][yStart - 1] = 4
        array[xStart][yStart] = 0 
        if xStart != xEnd or yStart-1 != yEnd:
            array[xEnd][yEnd] = 3
        else:
            Win()
        draw_maze(startx,starty,array)
        displaySurf.blit(startPic,(startx + xStart*cell_size,starty + (yStart-1)*cell_size))
        displaySurf.blit(endPic,(startx + xEnd*cell_size,starty + yEnd*cell_size))
        
    elif stable == 4 and yStart + 1 < h and (array[xStart][yStart + 1] == 0 or array[xStart][yStart + 1] == 3):
        print("oooooooooooooooo")
        array[xStart][yStart + 1] = 4 
        array[xStart][yStart] = 0
        if xStart != xEnd or yStart+1 != yEnd:
            array[xEnd][yEnd] = 3
        else:
            Win()
        draw_maze(startx,starty,array)
        displaySurf.blit(startPic,(startx + xStart*cell_size,starty + (yStart+1)*cell_size))
        displaySurf.blit(endPic,(startx + xEnd*cell_size,starty + yEnd*cell_size))
    pg.time.delay(100)

while(True):
    while(AiScreen):
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            if check_stable_button(border,border/2*2+50*1,buttonbarW-border,50):
                imgStart.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15)
                imgEnd.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15*4)
                check_button_DFS_maze = True
                check_user_play = False
            if check_stable_button(border,border/2*3+50*2,buttonbarW-border,50):
                imgStart.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15)
                imgEnd.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15*4)
                check_button_Hunt_and_Kill_maze = True
                check_user_play = False
            if check_stable_button(border,border/2*6+50*5,buttonbarW-border,50):
                check_button_dfs = True
            if check_stable_button(border,border/2*5+50*4,buttonbarW-border,50):
                check_button_bfs = True
            if check_stable_button(border,border/2*7+50*6,buttonbarW-border,50):
                check_button_Greedy = True
            if check_stable_button(border,border/2*8+50*7,buttonbarW-border,50):
                check_button_Astar = True
            if check_stable_button(border,border/2*8+50*8,buttonbarW-border,50):
                check_button_localsearch = True
            if check_stable_button(border+buttonbarW+mazeSize,border/2+50*3,140,40):
                AiScreen=False
                haveMap=False
                HomeScreen=True
                break
            if check_stable_button(border+buttonbarW+mazeSize,border/2+50*2,140,40):
                print("Choi")
                check_user_play = not check_user_play
            if check_stable_button(border+buttonbarW+mazeSize+10,border/2+50*5+40*1,120,30):
                if inhand==1:
                    inhand=0
                else:
                    inhand=1
            if check_stable_button(border+buttonbarW+mazeSize+10,border/2+50*5+40*2,120,30):
                if inhand==2:
                    inhand=0
                else:
                    inhand=2
            if check_stable_button(border+buttonbarW+mazeSize+10,border/2+50*5+40*3,120,30):
                for i in range(1,24):
                    for j in range(1,24):
                        if MAZE[i][j]==1: 
                            MAZE[i][j]=0
                draw_maze(startx,starty,MAZE)
                if checkCreateMap:
                    for i in range(1,24):
                        for j in range(1,24):
                            if MAZE[i][j]==1: 
                                MAZE[i][j]=0
                    draw_maze(startx,starty,MAZE)
            if check_stable_button(border+buttonbarW+mazeSize,border/2+50*4,20,20) and indexMap>0:
                if(indexMap>0):
                    indexMap=indexMap-1
            if check_stable_button(border+buttonbarW+mazeSize+50,border/2+50*4,20,20):
                if(indexMap<numOfMap):
                    indexMap=indexMap+1
            if check_stable_button(border+buttonbarW+mazeSize+70,border/2+50*4,70,20):
                if indexMap!=0:
                    imgStart.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15)
                    imgEnd.topleft =(border+buttonbarW+mazeSize+75+10,border/2+15*4)
                    MAZE=ReadMaze()

            if check_stable_button(border+buttonbarW+mazeSize+70,border/2+50*4+25,70,20):
                indexMap=numOfMap+1
                WriteMaze(MAZE)
            if event.type == pg.MOUSEBUTTONDOWN:
                if event.button == 1:
                    if imgStart.collidepoint(event.pos):
                        draggingStart = True  # Bắt đầu kéo
                        mouse_x, mouse_y = event.pos
                        offset_x = imgStart.x - mouse_x
                        offset_y = imgStart.y - mouse_y
                    if imgEnd.collidepoint(event.pos):
                        draggingEnd = True
                        mouse_x, mouse_y = event.pos
                        offset_x = imgEnd.x - mouse_x
                        offset_y = imgEnd.y - mouse_y
                    # print("++++++++++++++++")
                    # print(imgStart.x,imgStart.y)
                    # print(imgEnd.x,imgEnd.y)
                    beforexStart = imgStart.x
                    beforeyStart = imgStart.y
                    beforexEnd = imgEnd.x
                    beforeyEnd = imgEnd.y

            # Kiểm tra nếu thả chuột trái
            if event.type == pg.MOUSEBUTTONUP:
                if event.button == 1:  # Chuột trái
                    draggingStart = False  # Ngừng kéo
                    draggingEnd = False
                    if checkCreateStartEndDFS or checkCreateStartEndHuntAndKill:

                        kStart = int((imgStart.x - startx)//cell_size)
                        lStart = int((imgStart.y - starty)//cell_size)
                        mStart = int((beforexStart - startx)//cell_size)
                        nStart = int((beforeyStart - starty)//cell_size)
                        if kStart >=0 and kStart < 25 and lStart >= 0  and lStart < 25:
                            if MAZE[kStart][lStart] == 0:
                                MAZE[kStart][lStart] = 4
                                if mStart >= 0 and mStart <25 and nStart >= 0 and nStart < 25:
                                    MAZE[mStart][nStart] = 0
                            if MAZE[kStart][lStart] == 1:
                                if mStart >= 0 and mStart <25 and nStart >= 0 and nStart < 25:
                                    MAZE[mStart][nStart] = 4
                        else:
                            if mStart >= 0 and mStart <25 and nStart >= 0 and nStart < 25:
                                    MAZE[mStart][nStart] = 4

                        kEnd = int((imgEnd.x - startx)//cell_size)
                        lEnd = int((imgEnd.y - starty)//cell_size)
                        mEnd = int((beforexEnd - startx)//cell_size)
                        nEnd = int((beforeyEnd - starty)//cell_size)

                        if kEnd >= 0 and kEnd < 25 and lEnd >= 0 and lEnd < 25:
                            if MAZE[kEnd][lEnd] == 0:
                                    MAZE[kEnd][lEnd] = 3
                                    if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                                        MAZE[mEnd][nEnd] = 0
                            if MAZE[kEnd][lEnd] == 1:
                                    if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                                        MAZE[mEnd][nEnd] = 3
                        else:
                            if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                                    MAZE[mEnd][nEnd] = 3
                        

                        #print("dmmmmm")
                        draw_maze(startx, starty,MAZE)
                        
                        # kStart = int((imgStart.x - startx)//cell_size)
                        # lStart = int((imgStart.y - starty)//cell_size)
                        # mStart = int((beforexStart - startx)//cell_size)
                        # nStart = int((beforeyStart - starty)//cell_size)
                        
                        #print("llll",kStart,lStart)
                            
                            #print(MAZE[int((imgStart.x - startx)//cell_size)][int((imgStart.y - starty)//cell_size)])
                        if kStart >=0 and kStart < 25 and lStart >= 0  and lStart < 25:

                            # print("--------------------------------")
                            # print(imgStart.x,imgStart.y)
                            # print(startx,starty)
                            # print(int((imgStart.x - startx)//cell_size),int((imgStart.y - starty)//cell_size))
                            # print(MAZE[kStart][lStart])
                            # print("--------------------------------")

                            if MAZE[kStart][lStart] == 4:
                                # MAZE[kStart][lStart] = 4
                                #print("iiii")
                                checkCreateStartPoint = True
                                displaySurf.blit(startPic,(startx + kStart*cell_size,starty + lStart*cell_size))
                                imgStart.x = (startx + kStart*cell_size)
                                imgStart.y = (starty + lStart*cell_size)
                            if MAZE[kStart][lStart] == 1:
                                displaySurf.blit(startPic,(startx + mStart*cell_size,starty + nStart*cell_size))
                                imgStart.x = beforexStart
                                imgStart.y = beforeyStart
                                #draw_maze(startx, starty,MAZE)
                                #print("kkkkkk")
                        else: 
                            displaySurf.blit(startPic,(startx + mStart*cell_size,starty + nStart*cell_size))
                            imgStart.x = beforexStart
                            imgStart.y = beforeyStart
                                            
                        #print("kkkk",kEnd,lEnd)

                        if kEnd >= 0 and kEnd < 25 and lEnd >= 0 and lEnd < 25:

                            # print("++++++++++++++++++++++++++++++++++++++++++++++++")
                            # print(imgEnd.x,imgEnd.y)
                            # print(startx,starty)
                            # print(int((imgEnd.x - startx)//cell_size),int((imgEnd.y - starty)//cell_size))
                            # print(MAZE[kEnd][lEnd])
                            # print("++++++++++++++++++++++++++++++++++++++++++++++++")

                            
                            if MAZE[kEnd][lEnd] == 3:
                                # MAZE[kEnd][lEnd] = 3
                                # if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                                #     MAZE[mEnd][nEnd] = 0
                                displaySurf.blit(endPic,(startx + kEnd*cell_size,starty + lEnd*cell_size))
                                imgEnd.x = (startx + kEnd*cell_size)
                                imgEnd.y = (starty + lEnd*cell_size)
                                checkCreateEndPoint = True
                            if MAZE[kEnd][lEnd] == 1:
                                # if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                                #     MAZE[mEnd][nEnd] = 3
                                displaySurf.blit(endPic,(startx + mEnd*cell_size,starty + nEnd*cell_size))
                                imgEnd.x = beforexEnd
                                imgEnd.y = beforeyEnd
                        else:
                            # if mEnd >= 0 and mEnd <25 and nEnd >= 0 and nEnd < 25:
                            #         MAZE[mEnd][nEnd] = 3
                            displaySurf.blit(endPic,(startx + mEnd*cell_size,starty + nEnd*cell_size))
                            imgEnd.x = beforexEnd
                            imgEnd.y = beforeyEnd


                            # if MAZE[kEnd][lEnd] == 0:
                            #     displaySurf.blit(endPic,(startx + kEnd*cell_size,starty + lEnd*cell_size))
                            #     imgEnd.x = (startx + kEnd*cell_size)
                            #     imgEnd.y = (starty + lEnd*cell_size)
                            # else:
                            #     displaySurf.blit(endPic,(startx + mEnd*cell_size,starty + nEnd*cell_size))
                            #     imgEnd.x = beforexEnd
                            #     imgEnd.y = beforeyEnd
                                #draw_maze(startx, starty,MAZE)
                        # else:
                        #     displaySurf.blit(endPic,(startx + mEnd*cell_size,starty + nEnd*cell_size))
                        #     imgEnd.x = beforexEnd
                        #     imgEnd.y = beforeyEnd
            # Di chuyển biểu tượng khi kéo
            if event.type == pg.MOUSEMOTION:
                if draggingStart:
                    mouse_x, mouse_y = event.pos
                    imgStart.x = mouse_x + offset_x
                    imgStart.y = mouse_y + offset_y
                if draggingEnd:
                    mouse_x, mouse_y = event.pos
                    imgEnd.x = mouse_x + offset_x
                    imgEnd.y = mouse_y + offset_y
            if checkCreateMap and EditMaze(border/2+buttonbarW+20*1,border/2+20*1,20*23,20*23,MAZE):
                draw_maze(startx, starty,MAZE)
                kStart = int((imgStart.x - startx)//cell_size)
                lStart = int((imgStart.y - starty)//cell_size)
                kEnd = int((imgEnd.x - startx)//cell_size)
                lEnd = int((imgEnd.y - starty)//cell_size)
                displaySurf.blit(endPic,(startx + kEnd*cell_size,starty + lEnd*cell_size))
                displaySurf.blit(startPic,(startx + kStart*cell_size,starty + lStart*cell_size))
        #di chuyển khi chơi
        keys = pg.key.get_pressed()
        if keys[pg.K_LEFT] and check_user_play == True and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            fix= int((imgEnd.x - startx)//cell_size)
            fiy = int((imgEnd.y - starty)//cell_size)
            print("================================================")
            print(1,"leng",len(End),End,Start,fix,fiy)
            print("================================================")
            if Start[0] > 0 and Start[1] < w:
                print(Start,End)
                imgStart.x = (startx + (Start[0]-1)*cell_size)
                imgStart.y = (starty + Start[1]*cell_size)
                # print(int((imgStart.x - startx)//cell_size))
                # print(int((imgStart.y - starty)//cell_size))
                move(MAZE,Start[0],Start[1],fix,fiy,1)
            #handle(Begin,100,100,90,3,1)
            
            #print("mũi tên trái được nhấn")
        if keys[pg.K_RIGHT] and check_user_play == True and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            fix= int((imgEnd.x - startx)//cell_size)
            fiy = int((imgEnd.y - starty)//cell_size)
            print("================================================")
            print(2,"leng",len(End),End,Start,fix,fiy)
            print("================================================")
            if Start[0] > 0 and Start[1] < w:
                imgStart.x = (startx + (Start[0]+1)*cell_size)
                imgStart.y = (starty + Start[1]*cell_size)
                # print(Start,End)
                # print(int((imgStart.x - startx)//cell_size))
                # print(int((imgStart.y - starty)//cell_size))
                move(MAZE,Start[0],Start[1],fix,fiy,2)
            #print(2)
            #handle(Begin,100,100,90,3,2)
        if keys[pg.K_UP] and check_user_play == True and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            fix= int((imgEnd.x - startx)//cell_size)
            fiy = int((imgEnd.y - starty)//cell_size)
            print("================================================")
            print(3,"leng",len(End),End,Start,fix,fiy)
            print("================================================")
            if Start[0] > 0 and Start[1] < w:
                imgStart.x = (startx + Start[0]*cell_size)
                imgStart.y = (starty + (Start[1]-1)*cell_size)
                # print(Start,End)
                # print(int((imgStart.x - startx)//cell_size))
                # print(int((imgStart.y - starty)//cell_size))
                move(MAZE,Start[0],Start[1],fix,fiy,3)
            #print(3)
            #handle(Begin,100,100,90,3,3)
            
            #print(Begin)
        if keys[pg.K_DOWN] and check_user_play == True and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            fix= int((imgEnd.x - startx)//cell_size)
            fiy = int((imgEnd.y - starty)//cell_size)
            print("================================================")
            print(4,"leng",len(End),End,Start,fix,fiy)
            print("================================================")
            if Start[0] > 0 and Start[1] < w :
                imgStart.x = (startx + Start[0]*cell_size)
                imgStart.y = (starty + (Start[1]+1)*cell_size)
                
                # print(Start,End)
                # print(int((imgStart.x - startx)//cell_size))
                # print(int((imgStart.y - starty)//cell_size))
                
                move(MAZE,Start[0],Start[1],fix,fiy,4)
        
        drawAiScreen()
        if start:
            draw_maze(startx,starty,MAZE) 
            start=False
        if check_button_DFS_maze:
            #print(startx , randomstartx, starty, randomstarty)
            #(randomstartx,randomstarty)
            randomstartx = random.randint(1,24)
            randomstarty = random.randint(1,24)
            while randomstartx%2==0:
                randomstartx = random.randint(1,24)
            while randomstarty%2==0:
                randomstarty = random.randint(1,24)
            SaveStartx = randomstartx
            SaveStarty = randomstarty
            maze = [[1 for _ in range(w)] for _ in range(h)]  # 1: Tường, 0: Đường
            generate_maze(randomstartx,randomstarty,maze)
            MAZE = maze
            draw_maze(startx,starty,MAZE)
            checkCreateStartEndDFS = True
            checkCreateMap = True
            check_button_DFS_maze = False
            
        if check_button_Hunt_and_Kill_maze:
            n=int(500/cell_size)
            randomstartx = random.randint(0,n-1)
            randomstarty = random.randint(0,n-1)
            while randomstartx%2==0:
                randomstartx = random.randint(1,n-1)
            while randomstarty%2==0:
                randomstarty = random.randint(1,n-1)
            SaveStartx = randomstartx
            SaveStarty = randomstarty
            maze = [[0 for _ in range(n)] for _ in range(n)]
            createMaze(maze,n)
            huntAndKill(maze,SaveStartx,SaveStarty)
            #print(MAZE)
            draw_maze(startx,starty,maze)
            for i in range(w):
                for j in range(h):
                    if maze[i][j] == 2:
                        maze[i][j] = 0
            MAZE = maze
            checkCreateStartEndHuntAndKill = True
            checkCreateMap = True
            check_button_Hunt_and_Kill_maze = False
        if check_button_dfs and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            Maze =copy.deepcopy(MAZE)
            path = finddfs(Maze,Start,End)
            startX,startY = Start
            endx,endy = End
            Total = 0
            for i in range(w):
                for j in range(h):
                    if Maze[i][j] == 5:
                        Total += 1
            if path is None:
                drawLose()
                check_button_dfs =  False
            else:
                Moving(Maze,Start,End,path)
                saveRowsData([count,"DFS", f"({startX}, {startX})", f"({endx}, {endy})", Total, len(path)])
                check_button_dfs =  False
                drawWin()

        if check_button_bfs and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            Maze =copy.deepcopy(MAZE)
            path = findbfs(Maze,Start,End)
            startX,startY = Start
            endx,endy = End
            Total = 0
            for i in range(w):
                for j in range(h):
                    if Maze[i][j] == 5:
                        Total += 1
            if path is None:
                drawLose()
                check_button_bfs = False
            else:
                Moving(Maze,Start,End,path)
                saveRowsData([count,"BFS", f"({startX}, {startX})", f"({endx}, {endy})", Total, len(path)])
                check_button_bfs = False
                drawWin()

        #Geedy and Astar
        if check_button_Greedy and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:      
            Start,End = findStartEnd(MAZE)
            Maze =copy.deepcopy(MAZE)
            path = findGreedy(Maze,Start,End)
            startX,startY = Start
            endx,endy = End
            Total = 0
            for i in range(w):
                for j in range(h):
                    if Maze[i][j] == 5:
                        Total += 1
            if path is None:
                drawLose()
                check_button_Greedy =  False
            else:
                Moving(Maze,Start,End,path)
                saveRowsData([count,"Greedy", f"({startX}, {startX})", f"({endx}, {endy})", Total, len(path)])
                check_button_Greedy =  False
                drawWin()

        if check_button_Astar and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
            Start,End = findStartEnd(MAZE)
            Maze =copy.deepcopy(MAZE)
            path = findAstar(Maze,Start,End)
            startX,startY = Start
            endx,endy = End
            Total = 0
            for i in range(w):
                for j in range(h):
                    if Maze[i][j] == 5:
                        Total += 1
            if path is None:
                drawLose()
                check_button_Astar =  False
            else:
                Moving(Maze,Start,End,path)
                saveRowsData([count,"A*", f"({startX}, {startX})", f"({endx}, {endy})", Total, len(path)])
                check_button_Astar =  False
                drawWin()
        #local search
        if check_button_localsearch and checkCreateMap and checkCreateEndPoint and checkCreateStartPoint:
        #print("222")
            Start,End = findStartEnd(MAZE)
            Maze =copy.deepcopy(MAZE)
            path = findlocalsearch(Maze,Start,End)
            startX,startY = Start
            endx,endy = End
            Total = 0
            for i in range(w):
                for j in range(h):
                    if Maze[i][j] == 5:
                        Total += 1
            if path is None:
                drawLose()
                check_button_localsearch = False
            else:
                Moving(Maze,Start,End,path)
                saveRowsData([count,"Local search", f"({startX}, {startX})", f"({endx}, {endy})", Total, len(path)])
                check_button_localsearch = False
                drawWin()
        pg.display.update()

    while(HomeScreen):
        drawHomeScreen()
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            if check_stable_button(screenW/2-100,250,200,50):
                HomeScreen=False
                AiScreen=True
                haveMap=False
            if check_stable_button(screenW/2-100,325,200,50):
                HomeScreen=False
                AboutUsScreen=True
            if check_stable_button(screenW/2-100,400,200,50):
                HomeScreen=False
                HistoryScreen=True
        pg.display.update()

    while(AboutUsScreen):
        drawAboutUsScreen()
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            if check_stable_button(screenW/2-100,400,200,50):
                AboutUsScreen=False
                HomeScreen=True
        pg.display.update()
    while(WinScreen):
        drawWinScreen()
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            if check_stable_button(screenW/2-100,250,200,50):
                WinScreen=False
                AiScreen=True
        pg.display.update()
    while(HistoryScreen):
        drawHistoryScreen()
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            elif event.type == pg.MOUSEBUTTONDOWN:
                if event.button == 4:  # Lăn chuột lên
                    scroll_y = max(scroll_y - SCROLL_SPEED, 0)
                elif event.button == 5:  # Lăn chuột xuống
                    scroll_y = min(scroll_y + SCROLL_SPEED, tableHeight - screenH + 100)
            if check_stable_button(screenW/2-350,50,200,50):
                HomeScreen=True
                HistoryScreen=False
        if dataTable!=0:
            drawTable(displaySurf, dataTable, 40, 100,scroll_y, cell_width=140, cell_height=50)
        pg.display.update()