import pandas as pd
import pygame as pg
import sys
from openpyxl import Workbook, load_workbook

HistoryScreen = True
mazeSize=500
buttonbarW=250
barH=500
pointbarW=150
border=10
screenW=border+mazeSize+buttonbarW+pointbarW
screenH=border+barH
bg=pg.image.load("./img/bg.png")
bg=pg.transform.scale(bg,(screenW,screenH))
displaySurf = pg.display.set_mode((screenW, screenH))
pg.display.set_caption("Vẽ Table trong Pygame")
# Tên file Excel
file_name = "History Algorithm.xlsx"
# Biến scroll
scroll_y = 0  # Vị trí cuộn theo trục Y
SCROLL_SPEED = 20  # Tốc độ cuộn

white = (255, 255, 255)
red = (255, 0, 0)  
aqua = (0, 255,255) 
pg.init()
def drawText(x,y,w,h,content,size,color):
    font=pg.font.Font(None,size)
    text=font.render(content,True,color)
    textRect=text.get_rect()
    textBlockRect=pg.Rect(x,y,w,h)
    textRect.center=textBlockRect.center
    displaySurf.blit(text,textRect)

def drawHistoryScreen():
    displaySurf.blit(bg,(0,0))
   

# Kiểm tra nếu file chưa tồn tại, tạo mới
try:
    wb = load_workbook(file_name)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    # Thêm tiêu đề cột
    # Định nghĩa các tên cột
    columns = [
    "STT",
    "Algorithm",
    "Start position",
    "End position",
    "Total steps",
    "Result steps" 
    ]
    ws.append(columns)
dataTable = 0
tableWidth = 0
tableHeight = 0
wb.save(file_name) 
df = pd.read_excel(file_name)
lenTable = [df.columns.tolist()] + df.values.tolist()  # Chuyển thành danh sách 2D
count = len(lenTable) - 1
def saveRowsData(data):
    global dataTable,tableWidth ,tableHeight
    ws.append(data)
    wb.save(file_name)  # Lưu thay đổi sau mỗi lần ghi
    df = pd.read_excel(file_name)
    dataTable = [df.columns.tolist()] + df.values.tolist()  # Chuyển thành danh sách 2D
    #print(dataTable,"88888888")
    
    tableWidth = len(dataTable[0]) * 200
    tableHeight = len(dataTable) * 50
    

def drawTable(surface, data, start_x, start_y,scroll_y,cell_width, cell_height):
   # Font chữ
    font = pg.font.Font(None, 30)
    #print(data)
    drawText(450,50-scroll_y,50,50,"History Algorithm",50,white)
    for row_idx, row in enumerate(data):
        y = start_y + row_idx * cell_height - scroll_y
        if y + cell_height < 0 or y > screenH:  # Bỏ qua nếu ngoài màn hình
            continue
        for col_idx, cell in enumerate(row):
            x = start_x + col_idx * cell_width
            # Vẽ ô
            pg.draw.rect(surface, white, (x, y, cell_width, cell_height), 1)
            # Vẽ nội dung
            if col_idx == 4: 
                text_surface = font.render(str(cell), True, red)
            elif col_idx == 5:
                text_surface = font.render(str(cell), True, aqua)
            else:
                text_surface = font.render(str(cell), True, white)
            text_rect = text_surface.get_rect(center=(x + cell_width // 2, y + cell_height // 2))
            surface.blit(text_surface, text_rect)
     



saveRowsData([count,"A*", "(0, 0)", "(5, 5)", 15, 10])

#print(dataTable,"11111")
while(True):
    while(HistoryScreen):
        drawHistoryScreen()
        for event in pg.event.get():
            if event.type==pg.QUIT:
                    pg.quit()
                    sys.exit()
            elif event.type == pg.MOUSEBUTTONDOWN:
                if event.button == 4:  # Lăn chuột lên
                    scroll_y = max(scroll_y - SCROLL_SPEED, 0)
                elif event.button == 5:  # Lăn chuột xuống
                    scroll_y = min(scroll_y + SCROLL_SPEED, tableHeight - screenH + 100)
        drawTable(displaySurf, dataTable, 40, 100,scroll_y, cell_width=140, cell_height=50)
        pg.display.update()